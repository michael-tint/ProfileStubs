# -*- coding: utf-8 -*-
"""
Created on Mon Nov 25 11:04:34 2019

@author: TintM
"""

import xlrd 
import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
import time

def Name(profiles,data):
    for i in range(0,len(profiles['output'])):
        swap = str(profiles.loc[i,"Prog_Name"])
        profiles.loc[[i],["output"]] = str(profiles.loc[i,"output"]).replace("<name>",swap,1)
    return profiles


def Operators(profiles,data):
    data2=pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Operator"], aggfunc=np.sum)
    data2.reset_index(inplace=True)
    
    for i in range(0,len(profiles['output'])):
        
        data3 = data2[data2['Prog_Name'] == profiles.loc[i,"Prog_Name"]]
        data3.reset_index(inplace=True)
        
        if len(data3)==0:
            swap=""
        
        elif len(data3)==1:
            swap = data3.loc[0,'Operator']
            swap = 'the ' + swap 
        
        else:
            swap = str(len(data3))
            swap = swap + ' operators'               
        
        profiles.loc[[i],["output"]] = str(profiles.loc[i,"output"]).replace("<operators>",swap,1)
    
    return profiles

def Countries(profiles,data):
    data2=pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Country"], aggfunc=np.sum)
    data2.reset_index(inplace=True)
    
    for i in range(0,len(profiles['output'])):
        
        data3 = data2[data2['Prog_Name'] == profiles.loc[i,"Prog_Name"]]
        data3.reset_index(inplace=True)
        
        if len(data3)==0:
            swap= ""
        
        elif len(data3)==1:
            swap = data3.loc[0,'Country']
            swap = TheCountries(swap)
        elif len(data3)<6:
            swap = SeriesList(data3,'Country',' and')
            swap = TheCountries(swap)
        else:
            swap = str(len(data3))
            swap = swap + ' countries'
        
        profiles.loc[[i],["output"]] = str(profiles.loc[i,"output"]).replace("<countries>",swap,1)
    
    return profiles

def TheCountries(swap):  
        
    swap = swap.replace("USA","the United States")    
    swap = swap.replace("United Kingdom","the United Kingdom")
    swap = swap.replace("Netherlands","the Netherlands")
    swap = swap.replace("Philippines","the Philippines")
    swap = swap.replace("Czech Republic","the Czech Republic")
    swap = swap.replace("Dominican Republic","the Dominican Republic")
    swap = swap.replace("Maldives","the Maldives")
    swap = swap.replace("United Arab Emirates","the United Arab Emirates")
   
    return swap



def ISFandOrders(profiles,data):
    
    isfdata = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name"], aggfunc= "sum")

    ordersdata = data[data["Entry type"]=="New Build"]
    ordersdata = ordersdata[ordersdata["Entry certainty"].isin(["Contracted","Long Lead"])]
    ordersdata = pd.pivot_table(ordersdata, values="Total_Deliveries_Y0toY10", index=["Prog_Name"], aggfunc=np.sum)

    
    for i in range(0,len(profiles['output'])):
      
        isf = sum(isfdata['Current Fleet'])
        
        try:
            orders = sum(ordersdata['Total_Deliveries_Y0toY10'])
        except KeyError:
            orders = 0
        
        swap = " As of December 2019,"
        
        if isf == 0:
            if orders == 0:
                swap = swap + " it was in military service with"
            elif orders == 1:
                swap = swap + " there was one new delivery on contract with"
            else:
                swap = swap + " there were " + format(orders, ',') + " new deliveries on contract with"
        
        elif isf == 1:   
            swap = swap + " there was one in military service"
            if orders == 0:
                swap = swap + " with"
            elif orders == 1:
                swap = swap + " and one on order with"
            else:
                swap = swap + " and " + format(orders, ',') + " new deliveries on contract with"            
        else:
            swap = swap + " there were " + format(isf, ',') + " in military service"
            if orders == 0:
                swap = swap + " with"
            elif orders == 1:
                swap = swap + " and one on order with"
            else:
                swap = swap + " and " + format(orders, ',') + " new deliveries on contract with"             
    
        profiles.loc[[i],["output"]] = str(profiles.loc[i,"output"]).replace("<ISFandOrders>",swap,1)
    
    return profiles

def Orders(profiles,data):
    
    data2 = data[data["Entry type"]=="New Build"]
    data2 = data2[data2["Entry certainty"]=="Contracted"]
    data2 = pd.pivot_table(data2, values="Total_Deliveries_Y0toY10", index=["Prog_Name"], aggfunc=np.sum)
    
    for i in range(0,len(profiles['output'])):
        
        try:
            orders = data2.loc[profiles.loc[i,"Prog_Name"],"Total_Deliveries_Y0toY10"]
        except KeyError:
            order = 0
            
        if orders == 0:
            swap = ""
        
        elif orders == 1:
            swap = "and one new delivery on contract"
            
        else:
            swap = "and " + str(data2.loc[profiles.loc[i,"Prog_Name"],"Total_Deliveries_Y0toY10"]) + " new deliveries on contract"
        
        profiles.loc[[i],["output"]] = str(profiles.loc[i,"output"]).replace("<Orders>",swap,1)
    
    return profiles
    
def ISF(profiles,data):
    data2 = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name"], aggfunc=np.sum)
    
    for i in range(0,len(profiles['output'])):
        swap = "On December 1st, 2019"
        if data2.loc[profiles.loc[i,"Prog_Name"],"Current Fleet"] == 0:
            swap = swap + " none were in service"
        elif data2.loc[profiles.loc[i,"Prog_Name"],"Current Fleet"] == 1:
            swap = swap + " there was " + str(data2.loc[profiles.loc[i,"Prog_Name"],"Current Fleet"]) + " in service"
        else:
            swap = swap + " there were " + str(data2.loc[profiles.loc[i,"Prog_Name"],"Current Fleet"]) + " in service"
        
        profiles.loc[[i],["output"]] = str(profiles.loc[i,"output"]).replace("<ISF>",swap,1)
    
    return profiles

def Descriptor(profiles,data):
    for i in range(0,len(profiles['output'])):
        
        swap = str(profiles.loc[i,"Prog_Cat"])
        if swap == 'Maritime, C4ISR & Gunships':
            swap = '<mission>'        
        elif swap == 'Light 1-2 Seaters (Piston / Turboprop / Jet)':
            swap = 'light 1-2 seat aircraft'
            
        elif swap == 'Military Transports & Tankers':
            
            data2 = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Fixed Wing Weight"], aggfunc=np.sum)
            data2.reset_index(inplace=True)
            data2 = data2[data2['Prog_Name'] == profiles.loc[i,"Prog_Name"]]
            data2.reset_index(inplace=True)
            
            swap =  data2.loc[0,'Fixed Wing Weight'] + ' transport aircraft'
            swap = swap.lower()
            
        elif swap == 'Civil Propeller (3+ Seats)':
            swap = 'civil propeller aircraft'
        elif swap == 'UAVs':
            data2 = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","AircraftCat"], aggfunc=np.sum)
            data2.reset_index(inplace=True)
                           
            swap = SeriesList(data2,"AircraftCat",' or')
            swap = swap.replace("UAV","")
            swap = swap.replace("Group","")
            swap = "Group" + swap + "UAV"
            

        elif swap == 'Helicopters & Tiltrotors':
            swap = 'helicopter'           
        elif swap == 'Fighters':
            swap = 'fighter'
        elif swap == 'Commercial Jets':
            swap = 'commercial jet'
        elif swap == 'Bombers':
            swap = 'bomber'
        elif swap == 'Amphibians':
            swap = 'amphibious aircraft'
        elif swap == 'Business Jets':
            swap = 'business jet'
            
        profiles.loc[i,"output"] = str(profiles.loc[i,"output"]).replace("<descriptor>",swap,1)
    
    return profiles

def Typecount(profiles,data):
    data2 = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Type"], aggfunc="sum")
    data2.reset_index(inplace=True)

    data2 = pd.pivot_table(data2, values="Type", index=["Prog_Name"], aggfunc="count")

    for i in range(0,len(profiles['output'])):
        swap = data2.loc[profiles.loc[i,"Prog_Name"],"Type"]
        if swap == 1:
            swap = "one type"
        else:
            swap = str(swap) + " types"
        profiles.loc[i,"output"] = str(profiles.loc[i,"output"]).replace("<typecount>",swap,1)

    return profiles

def SeriesList(data,series,last):
    
    temp = data[series].str.cat(sep=', ', na_rep='?')
    temp = LastReplace(temp,",",last)
    return temp
    

def LastReplace(s, old, new):
    return (s[::-1].replace(old[::-1],new[::-1], 1))[::-1]

def TypeList(profiles,data):  
    data2 = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Type"], aggfunc='sum')
    data2.reset_index(inplace=True)
    
    for i in range(0,len(profiles['output'])):
        
        data3 = data2[data2['Prog_Name'] == profiles.loc[i,"Prog_Name"]]
        data3.reset_index(inplace=True)
        
        if len(data3)>20:
            swap = ""
        
        elif len(data3)==1:
            swap = data3.loc[0,'Type']
            swap = ', the ' + swap 
        
        else:
            swap = LastReplace(", the " + data3['Type'].str.cat(sep=', ', na_rep='?'),","," and")
                 
        profiles.loc[i,"output"] = str(profiles.loc[i,"output"]).replace("<typelist>",swap,1)
    
    return profiles

def OtherProfiles(profiles,data):
    for i in range(0,len(profiles['output'])):
        profiles2 = profiles[profiles['Family'] == profiles.loc[i,"Family"]]
        
        if len(profiles2)==1:
            swap = ""
        
        elif len(profiles2)==2:
            profiles2 = profiles2[profiles2['Prog_Name'] != profiles.loc[i,"Prog_Name"]]
            swap = "This family also includes the " + SeriesList(profiles2,'Prog_FullName',' and') + " profile."                
        else:            
            profiles2 = profiles2[profiles2['Prog_Name'] != profiles.loc[i,"Prog_Name"]]
            swap = "Other profiles in this family include the " + SeriesList(profiles2,'Prog_FullName',' and') + "."
            
        profiles.loc[i,"output"] = str(profiles.loc[i,"output"]).replace("<otherprofiles>",swap,1)
      
    return profiles

def EngineCount(data):

    
    count = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","NbrEngs"], aggfunc='sum')
    count.reset_index(inplace=True)
    count['NbrEngs'] = count['NbrEngs'].astype(str)        
        
    ecount = SeriesList(count,'NbrEngs',' or') + " "
    
    ecount = TextNumbers(ecount)
    
    return ecount

def EngineFamily(data):
        
    family = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","EngineFamily"], aggfunc='sum')
    family.reset_index(inplace=True)
            
    if len(family['EngineFamily'])==1:
        if family.loc[0,'EngineFamily'] == "Indeterminate":
            efamily = ""
        else:
            efamily = family.loc[0,'EngineFamily'] + " " 
    else:
        family = family[family['EngineFamily'] != "Indeterminate"]
        efamily = SeriesList(family,'EngineFamily',' or') + " "
        
    return efamily

def EngineType(data):
    propulsion = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Propulsion"], aggfunc='sum')
    propulsion.reset_index(inplace=True)
    
    etype = SeriesList(propulsion,'Propulsion',' or') + " "
    etype = etype.lower()
    
    return etype

def AllEngines(profiles,data):
    
    for i in range(0,len(profiles['output'])):        
        
        ecount = EngineCount(data[data['Prog_Name'] == profiles.loc[i,"Prog_Name"]])
        efamily = EngineFamily(data[data['Prog_Name'] == profiles.loc[i,"Prog_Name"]])
        etype = EngineType(data[data['Prog_Name'] == profiles.loc[i,"Prog_Name"]])
        
        swap = ecount + efamily + etype
        
        if ecount == "one ":
            swap = swap + "engine"
        else:
            swap = swap + "engines"
                
        profiles.loc[i,"output"] = str(profiles.loc[i,"output"]).replace("<allengines>",swap,1)   
        
    return profiles

def Output(data,file,sheet):
    book = load_workbook(file)
    book[sheet].delete_rows(1,5000)
    
    writer = pd.ExcelWriter(file, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, sheet,index=False) 

def TextNumbers(string):
    
    string = string.replace('1','one')
    string = string.replace('2','two')
    string = string.replace('3','three')
    string = string.replace('4','four')
       
    return string

def ExportData(table,output_file,sheetname):
    book = load_workbook(output_file)
    book[sheetname].delete_rows(1,5000)
    
    writer = pd.ExcelWriter(output_file, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    table.to_excel(writer, sheetname,index=False)
    
    writer.save()

def LineBreak(profiles):
    
    for i in range(0,len(profiles['output'])):   
        profiles.loc[i,"output"] = str(profiles.loc[i,"output"]).replace("<lb>","\n")
    return profiles


print('started')
start_time = time.time()

inputfile="raw data.xlsx"
outputfile = 'output.xlsx'

xlsx = pd.ExcelFile(inputfile)

data = pd.read_excel(xlsx,"data")

profiles = pd.read_excel(xlsx,"profiles")  
profiles['Prog_AWIN_Acct_ID'] = profiles[['Prog_AWIN_Acct_ID']].applymap("{:.0f}".format)



profiles['output'] = profiles['input']

profiles = Name(profiles,data)
profiles = Operators(profiles,data)
profiles = ISFandOrders(profiles,data)
profiles = Descriptor(profiles,data)
profiles = Countries(profiles,data)
profiles = AllEngines(profiles,data)
profiles = Typecount(profiles,data)
profiles = TypeList(profiles,data)
profiles = OtherProfiles(profiles,data)
profiles = LineBreak(profiles)

output = profiles[['Prog_AWIN_Acct_ID','Prog_FullName','Prog_Cat','output']]
outputcheck = output[['output','Prog_FullName']]
ExportData(output,outputfile,'output')    

print("V1: --- %s seconds ---" % (time.time() - start_time))
