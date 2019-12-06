# -*- coding: utf-8 -*-
"""
Created on Fri Dec  6 11:50:44 2019

@author: TintM
"""

# -*- coding: utf-8 -*-
"""
Created on Mon Nov 25 11:04:34 2019

@author: TintM
"""

import xlrd 
import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
import time


def Name(profiles):
    
    swap = str(profiles["Prog_Name"])
    
    return swap

def Descriptor(profiles,data):
 
    swap = str(profiles["Prog_Cat"])
    if swap == 'Maritime, C4ISR & Gunships':
        swap = '<mission>'        
    elif swap == 'Light 1-2 Seaters (Piston / Turboprop / Jet)':
        swap = 'light 1-2 seat aircraft'
        
    elif swap == 'Military Transports & Tankers':
        
        data = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Fixed Wing Weight"], aggfunc=np.sum)
        data.reset_index(inplace=True)
        
        swap = data.loc[0,'Fixed Wing Weight'] + ' transport aircraft'
        swap = swap.lower()
        
    elif swap == 'Civil Propeller (3+ Seats)':
        swap = 'civil propeller aircraft'
    elif swap == 'UAVs':
  
        data = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","AircraftCat"], aggfunc=np.sum)
        data.reset_index(inplace=True)
                       
        swap = SeriesList(data,"AircraftCat",' or')
        swap = swap.replace("UAV","")
        swap = swap.replace("Group","")
        swap = "Group" + swap + "UAV"

        
    elif swap == 'Helicopters & Tiltrotors':
        swap = 'helicopter'
    elif swap == 'Fighters':
        swap = 'fighter'
    elif swap == 'Commercial Jets':
        swap = 'commercial jet'
    elif swap == 'Bombers':
        swap = 'bomber'
    elif swap == 'Amphibians':
        swap = 'amphibious aircraft'
    elif swap == 'Business Jets':
        swap = 'business jet'

    return swap


def Operators(data):
    data=pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Operator"], aggfunc=np.sum)
    data.reset_index(inplace=True)
        
    if len(data)==0:
        swap=""
    
    elif len(data)==1:
        swap = data.loc[0,'Operator']
        swap = 'the ' + swap 
    
    else:
        swap = str(len(data))
        swap = swap + ' operators'               
        
    return swap

def Countries(data):

    data = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Country"], aggfunc=np.sum)
    data.reset_index(inplace=True)

    if len(data)==0:
        swap= ""
    
    elif len(data)==1:
        swap = data.loc[0,'Country']
        swap = TheCountries(swap)
    elif len(data)<6:
        swap = SeriesList(data,'Country',' and')
        swap = TheCountries(swap)
    else:
        swap = str(len(data))
        swap = swap + ' countries'
    
    return swap


def ISFandOrders(profiles,data):
    
    isf = sum(data['Current Fleet'])
    orders = sum(data['Total_Deliveries_Y0toY10'])
    
    swap = " As of December 2019,"
    
    if isf == 0:
        if orders == 0:
            swap = swap + " it was in military service with"
        elif orders == 1:
            swap = swap + " there was one new delivery on contract with"
        else:
            swap = swap + " there were " + format(orders, ',') + " new deliveries on contract with"
    
    elif isf == 1:   
        swap = swap + " there was one in military service"
        if orders == 0:
            swap = swap + " with"
        elif orders == 1:
            swap = swap + " and one on order with"
        else:
            swap = swap + " and " + format(orders, ',') + " new deliveries on contract with"            
    else:
        swap = swap + " there were " + format(isf, ',') + " in military service"
        if orders == 0:
            swap = swap + " with"
        elif orders == 1:
            swap = swap + " and one on order with"
        else:
            swap = swap + " and " + format(orders, ',') + " new deliveries on contract with"             

    return swap


def TypeCount(profiles,data):
    data = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Type"], aggfunc="sum")
    data.reset_index(inplace=True)

    swap = len(data)

    if swap == 1:
        swap = "one type"
    else:
        swap = str(swap) + " types"

    return swap

def TypeList(profiles,data):  
    data = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Type"], aggfunc='sum')
    data.reset_index(inplace=True)    
    
    if len(data)>20:
        swap = ""
    
    elif len(data)==1:
        swap = data.loc[0,'Type']
        swap = ', the ' + swap 
    
    else:
        swap = SwapLast(", the " + data['Type'].str.cat(sep=', ', na_rep='?'),","," and")
                 
    return swap

def OtherProfiles(profiles,current):

    if len(profiles)==1:
        swap = ""
    
    elif len(profiles)==2:
        profiles = profiles[profiles['Prog_Name'] != current]
        swap = "This family also includes the " + SeriesList(profiles,'Prog_FullName'," and") + " profile."                
    else:            
        profiles = profiles[profiles['Prog_Name'] != current]
        swap = "Other profiles in this family include the " + SeriesList(profiles,'Prog_FullName',' and') + "."
      
    return swap

def EngineCount(data):
    
    count = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","NbrEngs"], aggfunc='sum')
    count.reset_index(inplace=True)
    count['NbrEngs'] = count['NbrEngs'].astype(str)        
        
    ecount = SeriesList(count,'NbrEngs'," or") + " "
    ecount = TextNumbers(ecount)
    
    return ecount

def EngineFamily(data):
        
    family = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","EngineFamily"], aggfunc='sum')
    family.reset_index(inplace=True)
            
    if len(family['EngineFamily'])==1:
        if family.loc[0,'EngineFamily'] == "Indeterminate":
            efamily = ""
        else:
            efamily = family.loc[0,'EngineFamily'] + " " 
    else:
        family = family[family['EngineFamily'] != "Indeterminate"]
        efamily = SeriesList(family,'EngineFamily',' or') + " "
    
    return efamily

def EngineType(data):
    propulsion = pd.pivot_table(data, values="Current Fleet", index=["Prog_Name","Propulsion"], aggfunc='sum')
    propulsion.reset_index(inplace=True)
    
    etype = SeriesList(propulsion,'Propulsion',' or') + " "
    etype = etype.lower()
    
    return etype

def AllEngines(profiles,data):
    
        
    ecount = EngineCount(data)
    efamily = EngineFamily(data)
    etype = EngineType(data)
    
    swap = ecount + efamily + etype
    
    if ecount == "one ":
        swap = swap + "engine"
    else:
        swap = swap + "engines"
            
    return swap

def TheCountries(swap):  
        
    swap = swap.replace("USA","the United States")    
    swap = swap.replace("United Kingdom","the United Kingdom")
    swap = swap.replace("Netherlands","the Netherlands")
    swap = swap.replace("Philippines","the Philippines")
    swap = swap.replace("Czech Republic","the Czech Republic")
    swap = swap.replace("Dominican Republic","the Dominican Republic")
    swap = swap.replace("Maldives","the Maldives")
    swap = swap.replace("United Arab Emirates","the United Arab Emirates")
   
    return swap

def SeriesList(data,series,final):
    
    temp = data[series].str.cat(sep=', ', na_rep='?')
    temp = SwapLast(temp,",",final)
    return temp
    
def SwapLast(s, old, new):
    return (s[::-1].replace(old[::-1],new[::-1], 1))[::-1]

def Output(data,file,sheet):
    book = load_workbook(file)
    book[sheet].delete_rows(1,5000)
    
    writer = pd.ExcelWriter(file, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    data.to_excel(writer, sheet,index=False) 

def TextNumbers(string):
    
    string = string.replace('1','one')
    string = string.replace('2','two')
    string = string.replace('3','three')
    string = string.replace('4','four')
       
    return string

def ExportData(table,output_file,sheetname):
    book = load_workbook(output_file)
    book[sheetname].delete_rows(1,5000)
    
    writer = pd.ExcelWriter(output_file, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    table.to_excel(writer, sheetname,index=False)
    
    writer.save()

def SwapAll(text,old,new):
    
    text = text.replace(old, new)   
    return text   

print("started")
start_time = time.time()

inputfile="raw data.xlsx"
outputfile = 'output2.xlsx'

xlsx = pd.ExcelFile(inputfile)

data = pd.read_excel(xlsx,"data")

profiles = pd.read_excel(xlsx,"profiles")  
profiles['Prog_AWIN_Acct_ID'] = profiles[['Prog_AWIN_Acct_ID']].applymap("{:.0f}".format)
profiles['output'] = profiles['input']


for i in range(0,len(profiles['output'])):      
    
    currentprofile = profiles.loc[i,:]
    currentdata = data[data["Prog_Name"]==currentprofile["Prog_Name"]]
    
    profiles.loc[i,"output"] = profiles.loc[i,"output"].replace("<name>",Name(currentprofile))
    profiles.loc[i,"output"] = profiles.loc[i,"output"].replace("<descriptor>",Descriptor(currentprofile,currentdata))
    profiles.loc[i,"output"] = profiles.loc[i,"output"].replace("<operators>",Operators(currentdata))
    profiles.loc[i,"output"] = profiles.loc[i,"output"].replace("<countries>",Countries(currentdata))
    profiles.loc[i,"output"] = profiles.loc[i,"output"].replace("<ISFandOrders>",ISFandOrders(currentprofile,currentdata))
    profiles.loc[i,"output"] = profiles.loc[i,"output"].replace("<allengines>",AllEngines(currentprofile,currentdata))
    profiles.loc[i,"output"] = profiles.loc[i,"output"].replace("<typecount>",TypeCount(currentprofile,currentdata))
    profiles.loc[i,"output"] = profiles.loc[i,"output"].replace("<typelist>",TypeList(currentprofile,currentdata))
    profiles.loc[i,"output"] = profiles.loc[i,"output"].replace("<otherprofiles>",OtherProfiles(profiles[profiles['Family'] == currentprofile["Family"]],currentprofile['Prog_Name']))
    profiles.loc[i,"output"] = profiles.loc[i,"output"].replace("<lb>","")
    profiles.loc[i,"output"] = profiles.loc[i,"output"].replace("  "," ")

output = profiles[['Prog_AWIN_Acct_ID','Prog_FullName','Prog_Cat','output']]
outputcheck = output[['output','Prog_FullName']]
ExportData(output,outputfile,'output')    


print("V2: --- %s seconds ---" % (time.time() - start_time))
